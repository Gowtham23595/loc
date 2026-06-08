#!/usr/bin/env python3
"""
Create an Excel report of executable-ish LOC per C/C++ API.

Expected project shape:

    project_root/
      module1/
        src/
        inc/
      module2/
        src/
        inc/

Usage:

    python api_loc_report.py --root . --out api_loc_report.xlsx

    python api_loc_report.py --root1 old_workspace --root2 new_workspace --out api_loc_comparison.xlsx

    python api_loc_report.py --root1 old_workspace --root2 new_workspace --exclude-dir visualisation --out api_loc_comparison.xlsx

Install dependency if needed:

    python -m pip install openpyxl

Notes:
    This is a practical parser, not a full C++ compiler frontend. It handles
    normal C/C++ free functions, qualified methods, and many inline header
    methods. It ignores declarations/prototypes and counts meaningful
    non-comment lines inside each detected function body.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path


SOURCE_EXTENSIONS = {".c", ".cpp", ".h", ".hpp"}
MODULE_DIRS = ("src", "inc", "include")
DEFAULT_REPORT_NAME = "api_loc_report.xlsx"

CONTROL_KEYWORDS = {
    "if",
    "for",
    "while",
    "switch",
    "catch",
    "else",
    "do",
    "try",
    "case",
    "return",
    "sizeof",
}

NON_API_KEYWORDS = {
    "class",
    "struct",
    "union",
    "enum",
    "namespace",
}


@dataclass(frozen=True)
class ApiLoc:
    module: str
    area: str
    file: str
    api: str
    start_line: int
    end_line: int
    loc: int


@dataclass(frozen=True)
class ScopeTotals:
    files: int
    apis: int
    loc: int


def mask_comments_and_strings(source: str) -> str:
    """Replace comments/strings/chars with spaces while preserving newlines."""
    result: list[str] = []
    i = 0
    n = len(source)
    state = "code"

    while i < n:
        ch = source[i]
        nxt = source[i + 1] if i + 1 < n else ""

        if state == "code":
            if ch == "/" and nxt == "/":
                result.extend((" ", " "))
                i += 2
                state = "line_comment"
            elif ch == "/" and nxt == "*":
                result.extend((" ", " "))
                i += 2
                state = "block_comment"
            elif ch == '"':
                result.append(" ")
                i += 1
                state = "string"
            elif ch == "'":
                result.append(" ")
                i += 1
                state = "char"
            else:
                result.append(ch)
                i += 1
        elif state == "line_comment":
            if ch == "\n":
                result.append("\n")
                state = "code"
            else:
                result.append(" ")
            i += 1
        elif state == "block_comment":
            if ch == "*" and nxt == "/":
                result.extend((" ", " "))
                i += 2
                state = "code"
            else:
                result.append("\n" if ch == "\n" else " ")
                i += 1
        elif state == "string":
            if ch == "\\" and i + 1 < n:
                result.extend((" ", " "))
                i += 2
            elif ch == '"':
                result.append(" ")
                i += 1
                state = "code"
            else:
                result.append("\n" if ch == "\n" else " ")
                i += 1
        elif state == "char":
            if ch == "\\" and i + 1 < n:
                result.extend((" ", " "))
                i += 2
            elif ch == "'":
                result.append(" ")
                i += 1
                state = "code"
            else:
                result.append("\n" if ch == "\n" else " ")
                i += 1

    return "".join(result)


def line_number_at(source: str, index: int) -> int:
    return source.count("\n", 0, index) + 1


def find_matching_brace(masked: str, open_index: int) -> int | None:
    depth = 0
    for i in range(open_index, len(masked)):
        if masked[i] == "{":
            depth += 1
        elif masked[i] == "}":
            depth -= 1
            if depth == 0:
                return i
    return None


def find_matching_open_paren(text: str, close_index: int) -> int | None:
    depth = 0
    for i in range(close_index, -1, -1):
        if text[i] == ")":
            depth += 1
        elif text[i] == "(":
            depth -= 1
            if depth == 0:
                return i
    return None


def signature_fragment_before(masked: str, brace_index: int) -> str:
    """Return probable function signature text immediately before a body brace."""
    start = max(0, brace_index - 3000)
    fragment = masked[start:brace_index]

    # Keep only the text after the last statement/body boundary. This keeps
    # multiline signatures while dropping previous declarations.
    last_boundary = -1
    for token in (";", "{", "}"):
        pos = fragment.rfind(token)
        if pos > last_boundary:
            last_boundary = pos
    if last_boundary >= 0:
        fragment = fragment[last_boundary + 1 :]

    return fragment.strip()


def extract_api_name(signature: str) -> str | None:
    signature = re.sub(r"\s+", " ", signature).strip()
    if not signature or "(" not in signature or ")" not in signature:
        return None

    close_paren = signature.rfind(")")
    open_paren = find_matching_open_paren(signature, close_paren)
    if open_paren is None:
        return None

    before_paren = signature[:open_paren].strip()
    after_paren = signature[close_paren + 1 :].strip()

    # Exclude class/namespace/enum/etc. and control blocks.
    first_word = re.match(r"([A-Za-z_]\w*)", before_paren)
    if first_word and first_word.group(1) in NON_API_KEYWORDS:
        return None

    name_match = re.search(
        r"(operator\s*(?:[^\s(]+|\(\)|\[\])|(?:[A-Za-z_~]\w*::)*[A-Za-z_~]\w*)\s*$",
        before_paren,
    )
    if not name_match:
        return None

    api_name = re.sub(r"\s+", "", name_match.group(1))
    simple_name = api_name.split("::")[-1].replace("~", "")
    if simple_name in CONTROL_KEYWORDS or simple_name in NON_API_KEYWORDS:
        return None

    # These usually indicate a declaration-like macro or unusual non-function body.
    disallowed_after = ("=",)
    if after_paren.startswith(disallowed_after):
        return None

    return api_name


def looks_like_function_signature(signature: str) -> str | None:
    api_name = extract_api_name(signature)
    if not api_name:
        return None

    compact = re.sub(r"\s+", " ", signature).strip()
    if compact.startswith("#"):
        return None
    if re.search(r"\btypedef\b", compact):
        return None
    if re.search(r"\busing\b.*=", compact):
        return None
    if re.search(r"\b(?:if|for|while|switch|catch)\s*\(", compact):
        return None
    if re.search(r"\[\s*[^\]]*\]\s*\([^)]*\)\s*$", compact):
        return None

    return api_name


def is_executable_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith("#"):
        return False
    if re.fullmatch(r"[{};,\s]+", stripped):
        return False
    if re.fullmatch(r"(public|private|protected)\s*:", stripped):
        return False
    if re.fullmatch(r"(case\b.*|default)\s*:", stripped):
        return False
    if re.fullmatch(r"[A-Za-z_]\w*\s*:", stripped):
        return False
    if re.fullmatch(r"(else|try|do)\s*\{?", stripped):
        return False
    return True


def executable_loc_in_body(masked_source: str, open_brace: int, close_brace: int) -> int:
    body = masked_source[open_brace + 1 : close_brace]
    return sum(1 for line in body.splitlines() if is_executable_line(line))


def parse_apis_in_file(module: str, area: str, root: Path, path: Path) -> list[ApiLoc]:
    try:
        source = path.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        source = path.read_text(encoding="utf-8", errors="replace")

    masked = mask_comments_and_strings(source)
    apis: list[ApiLoc] = []
    i = 0

    while i < len(masked):
        if masked[i] != "{":
            i += 1
            continue

        signature = signature_fragment_before(masked, i)
        api_name = looks_like_function_signature(signature)
        if not api_name:
            i += 1
            continue

        close = find_matching_brace(masked, i)
        if close is None:
            i += 1
            continue

        rel_file = str(path.relative_to(root))
        apis.append(
            ApiLoc(
                module=module,
                area=area,
                file=rel_file,
                api=api_name,
                start_line=line_number_at(masked, i),
                end_line=line_number_at(masked, close),
                loc=executable_loc_in_body(masked, i, close),
            )
        )

        # Skip nested scopes/lambdas inside this API body.
        i = close + 1

    return apis


def discover_module_roots(root: Path) -> list[tuple[str, Path]]:
    modules: list[tuple[str, Path]] = []

    if any((root / name).is_dir() for name in MODULE_DIRS):
        modules.append((root.name, root))

    for child in sorted(root.iterdir()):
        if not child.is_dir() or child.name.startswith("."):
            continue
        if any((child / name).is_dir() for name in MODULE_DIRS):
            modules.append((child.name, child))

    return modules


def is_excluded_path(path: Path, scan_root: Path, exclude_dirs: set[str]) -> bool:
    if not exclude_dirs:
        return False

    try:
        relative_parts = path.relative_to(scan_root).parts
    except ValueError:
        relative_parts = path.parts

    return any(part.lower() in exclude_dirs for part in relative_parts)


def collect_api_locs(root: Path, exclude_dirs: set[str] | None = None) -> list[ApiLoc]:
    exclude_dirs = exclude_dirs or set()
    rows: list[ApiLoc] = []
    for module_name, module_path in discover_module_roots(root):
        for area in MODULE_DIRS:
            area_path = module_path / area
            if not area_path.is_dir():
                continue
            for path in sorted(area_path.rglob("*")):
                if is_excluded_path(path, area_path, exclude_dirs):
                    continue
                if path.is_file() and path.suffix.lower() in SOURCE_EXTENSIONS:
                    rows.extend(parse_apis_in_file(module_name, area, root, path))

    return sorted(rows, key=lambda row: (row.module, row.file, row.start_line, row.api))


def load_openpyxl():
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "Missing dependency: openpyxl\n"
            "Install it with: python -m pip install openpyxl"
        ) from exc
    return openpyxl


def write_excel(rows: list[ApiLoc], out_path: Path) -> None:
    openpyxl = load_openpyxl()

    wb = openpyxl.Workbook()
    detail = wb.active
    populate_detail_sheet(detail, "API_LOC", rows, "ApiLocTable")
    populate_summary_sheet(wb.create_sheet("Summary"), rows)
    style_workbook(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def write_comparison_excel(
    rows1: list[ApiLoc],
    rows2: list[ApiLoc],
    root1: Path,
    root2: Path,
    out_path: Path,
) -> None:
    openpyxl = load_openpyxl()

    wb = openpyxl.Workbook()
    populate_detail_sheet(wb.active, "Workspace1_API_LOC", rows1, "Workspace1ApiLocTable")
    populate_summary_sheet(wb.create_sheet("Workspace1_Summary"), rows1)
    populate_detail_sheet(wb.create_sheet("Workspace2_API_LOC"), "Workspace2_API_LOC", rows2, "Workspace2ApiLocTable")
    populate_summary_sheet(wb.create_sheet("Workspace2_Summary"), rows2)
    populate_comparison_sheet(wb.create_sheet("Comparison"), rows1, rows2, root1, root2)
    style_workbook(wb)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def populate_detail_sheet(sheet, title: str, rows: list[ApiLoc], table_name: str) -> None:
    from openpyxl.worksheet.table import Table, TableStyleInfo

    sheet.title = title
    headers = ["Module Name", "Area", "File Name", "API Name", "Start Line", "End Line", "LOC"]
    sheet.append(headers)
    for row in rows:
        sheet.append([row.module, row.area, row.file, row.api, row.start_line, row.end_line, row.loc])

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, [22, 10, 55, 35, 12, 12, 10])

    if sheet.max_row > 1:
        table = Table(displayName=table_name, ref=f"A1:G{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def populate_summary_sheet(sheet, rows: list[ApiLoc]) -> None:
    sheet.append(["Module Name", "Files", "APIs", "Total LOC", "Average LOC/API", "Max LOC/API"])

    module_files: dict[str, set[str]] = defaultdict(set)
    module_locs: dict[str, list[int]] = defaultdict(list)
    for row in rows:
        module_files[row.module].add(row.file)
        module_locs[row.module].append(row.loc)

    for module in sorted(module_locs):
        locs = module_locs[module]
        sheet.append(
            [
                module,
                len(module_files[module]),
                len(locs),
                sum(locs),
                round(sum(locs) / len(locs), 2) if locs else 0,
                max(locs) if locs else 0,
            ]
        )

    total_apis = len(rows)
    total_loc = sum(row.loc for row in rows)
    sheet.append([])
    sheet.append(
        [
            "Grand Total",
            len({row.file for row in rows}),
            total_apis,
            total_loc,
            round(total_loc / total_apis, 2) if total_apis else 0,
            max((row.loc for row in rows), default=0),
        ]
    )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, [22, 12, 12, 14, 18, 16])


def module_totals(rows: list[ApiLoc]) -> dict[str, ScopeTotals]:
    grouped: dict[str, list[ApiLoc]] = defaultdict(list)
    for row in rows:
        grouped[row.module].append(row)
    return {
        module: ScopeTotals(
            files=len({row.file for row in items}),
            apis=len(items),
            loc=sum(row.loc for row in items),
        )
        for module, items in grouped.items()
    }


def file_totals(rows: list[ApiLoc]) -> dict[tuple[str, str], ScopeTotals]:
    grouped: dict[tuple[str, str], list[ApiLoc]] = defaultdict(list)
    for row in rows:
        grouped[(row.module, row.file)].append(row)
    return {
        key: ScopeTotals(
            files=1,
            apis=len(items),
            loc=sum(row.loc for row in items),
        )
        for key, items in grouped.items()
    }


def api_map(rows: list[ApiLoc]) -> dict[tuple[str, str, str], ApiLoc]:
    result: dict[tuple[str, str, str], ApiLoc] = {}
    duplicates: dict[tuple[str, str, str], int] = defaultdict(int)

    for row in rows:
        key = (row.module, row.file, row.api)
        duplicates[key] += 1
        if duplicates[key] == 1:
            result[key] = row
        else:
            # Keep overloaded/repeated names visible instead of overwriting them.
            result[(row.module, row.file, f"{row.api}#{duplicates[key]}")] = row

    return result


def compare_status(present1: bool, present2: bool, loc1: int | None, loc2: int | None, unit: str) -> str:
    if present1 and not present2:
        return f"Missing {unit} in Workspace 2"
    if present2 and not present1:
        return f"New {unit} in Workspace 2"
    if loc1 != loc2:
        return "LOC Changed"
    return "Unchanged"


def populate_comparison_sheet(
    sheet,
    rows1: list[ApiLoc],
    rows2: list[ApiLoc],
    root1: Path,
    root2: Path,
) -> None:
    headers = [
        "Level",
        "Status",
        "Module Name",
        "File Name",
        "API Name",
        "Workspace 1 LOC",
        "Workspace 2 LOC",
        "LOC Delta",
        "Workspace 1 Files",
        "Workspace 2 Files",
        "Workspace 1 APIs",
        "Workspace 2 APIs",
        "Workspace 1 Start Line",
        "Workspace 2 Start Line",
        "Workspace 1 Root",
        "Workspace 2 Root",
    ]
    sheet.append(headers)

    totals1 = module_totals(rows1)
    totals2 = module_totals(rows2)
    for module in sorted(set(totals1) | set(totals2)):
        left = totals1.get(module)
        right = totals2.get(module)
        status = compare_status(left is not None, right is not None, left.loc if left else None, right.loc if right else None, "Module")
        sheet.append(
            [
                "Module",
                status,
                module,
                "",
                "",
                left.loc if left else "",
                right.loc if right else "",
                loc_delta(left.loc if left else None, right.loc if right else None),
                left.files if left else "",
                right.files if right else "",
                left.apis if left else "",
                right.apis if right else "",
                "",
                "",
                str(root1),
                str(root2),
            ]
        )

    files1 = file_totals(rows1)
    files2 = file_totals(rows2)
    for module, file_name in sorted(set(files1) | set(files2)):
        left = files1.get((module, file_name))
        right = files2.get((module, file_name))
        status = compare_status(left is not None, right is not None, left.loc if left else None, right.loc if right else None, "File")
        sheet.append(
            [
                "File",
                status,
                module,
                file_name,
                "",
                left.loc if left else "",
                right.loc if right else "",
                loc_delta(left.loc if left else None, right.loc if right else None),
                left.files if left else "",
                right.files if right else "",
                left.apis if left else "",
                right.apis if right else "",
                "",
                "",
                str(root1),
                str(root2),
            ]
        )

    apis1 = api_map(rows1)
    apis2 = api_map(rows2)
    for module, file_name, api_name in sorted(set(apis1) | set(apis2)):
        left = apis1.get((module, file_name, api_name))
        right = apis2.get((module, file_name, api_name))
        status = compare_status(left is not None, right is not None, left.loc if left else None, right.loc if right else None, "API")
        sheet.append(
            [
                "API",
                status,
                module,
                file_name,
                api_name,
                left.loc if left else "",
                right.loc if right else "",
                loc_delta(left.loc if left else None, right.loc if right else None),
                "",
                "",
                "",
                "",
                left.start_line if left else "",
                right.start_line if right else "",
                str(root1),
                str(root2),
            ]
        )

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    set_widths(sheet, [12, 30, 22, 55, 35, 16, 16, 12, 18, 18, 18, 18, 20, 20, 45, 45])


def loc_delta(loc1: int | None, loc2: int | None) -> int | str:
    if loc1 is None or loc2 is None:
        return ""
    return loc2 - loc1


def style_workbook(wb) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")

        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header_row(sheet) -> None:
    from openpyxl.styles import Font, PatternFill

    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")


def set_widths(sheet, widths: list[int]) -> None:
    for idx, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + idx)].width = width


def parse_args(argv: list[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate C/C++ API LOC Excel report.")
    parser.add_argument("--root", default=".", help="Project root containing module folders. Used for single-workspace mode.")
    parser.add_argument(
        "--root1",
        "--old",
        "--workspace1",
        dest="root1",
        help="First workspace root for comparison mode.",
    )
    parser.add_argument(
        "--root2",
        "--new",
        "--workspace2",
        dest="root2",
        help="Second workspace root for comparison mode.",
    )
    parser.add_argument(
        "--out",
        default=DEFAULT_REPORT_NAME,
        help="Output .xlsx path. If an existing folder is provided, api_loc_report.xlsx is created inside it.",
    )
    parser.add_argument(
        "--exclude-dir",
        action="append",
        default=[],
        help=(
            "Folder name to exclude anywhere below src/inc/include. "
            "Can be repeated or comma-separated, for example: "
            "--exclude-dir visualisation --exclude-dir generated,temp"
        ),
    )
    return parser.parse_args(argv)


def parse_exclude_dirs(values: list[str]) -> set[str]:
    exclude_dirs: set[str] = set()
    for value in values:
        for item in value.split(","):
            name = item.strip().strip("/\\")
            if name:
                exclude_dirs.add(name.lower())
    return exclude_dirs


def resolve_output_path(out_arg: str) -> Path:
    out_path = Path(out_arg).resolve()

    if out_path.exists() and out_path.is_dir():
        return out_path / DEFAULT_REPORT_NAME

    if out_path.suffix == "":
        return out_path.with_suffix(".xlsx")

    if out_path.suffix.lower() != ".xlsx":
        return out_path.with_suffix(".xlsx")

    return out_path


def main(argv: list[str]) -> int:
    args = parse_args(argv)
    out_path = resolve_output_path(args.out)
    exclude_dirs = parse_exclude_dirs(args.exclude_dir)

    comparison_mode = args.root1 is not None or args.root2 is not None

    if comparison_mode:
        if not args.root1 or not args.root2:
            print("Comparison mode requires both --root1 and --root2.", file=sys.stderr)
            return 2

        root1 = Path(args.root1).resolve()
        root2 = Path(args.root2).resolve()
        if not root1.exists():
            print(f"Workspace 1 root path does not exist: {root1}", file=sys.stderr)
            return 2
        if not root2.exists():
            print(f"Workspace 2 root path does not exist: {root2}", file=sys.stderr)
            return 2

        rows1 = collect_api_locs(root1, exclude_dirs)
        rows2 = collect_api_locs(root2, exclude_dirs)
        write_comparison_excel(rows1, rows2, root1, root2, out_path)

        print(f"Scanned workspace 1: {root1}")
        print(f"Workspace 1 APIs found: {len(rows1)}")
        print(f"Scanned workspace 2: {root2}")
        print(f"Workspace 2 APIs found: {len(rows2)}")
        if exclude_dirs:
            print(f"Excluded folders: {', '.join(sorted(exclude_dirs))}")
        print(f"Excel comparison report: {out_path}")
        return 0

    root = Path(args.root).resolve()
    if not root.exists():
        print(f"Root path does not exist: {root}", file=sys.stderr)
        return 2
    rows = collect_api_locs(root, exclude_dirs)
    write_excel(rows, out_path)

    print(f"Scanned root: {root}")
    print(f"APIs found: {len(rows)}")
    if exclude_dirs:
        print(f"Excluded folders: {', '.join(sorted(exclude_dirs))}")
    print(f"Excel report: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main(sys.argv[1:]))
